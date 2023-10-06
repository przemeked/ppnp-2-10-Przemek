import openpyxl

workbook = openpyxl.load_workbook('sales.xlsx')
worksheet = workbook.active

print(worksheet)

lista = []

for i in worksheet:
    print(i)

for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        lista.append(col[i].value)
        print(col[i].value)

print(lista)

print(lista[0])
print(lista[0:3]) 
